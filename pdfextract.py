''' Program to search keywords in PDFs and save the corresponding sentence
    to Excel Worksheet.  '''
import os
import glob
import re
import PyPDF2 as p2
from openpyxl import Workbook, load_workbook

folder_path = r"F:\Yashraj\yes\prac"
workbook = load_workbook(filename="Pkeyword.xlsx")
sheet = workbook.active

wb = Workbook()
ws = wb.active

ws['A1'] = 'Document Name'
ws['B1'] = 'Keyword'
ws['C1'] = 'Page No'
ws['D1'] = 'Content'

count = 1

def updating(word, password=''):
    global count
    for filename in glob.glob(os.path.join(folder_path, '*.pdf')):
        pdf_file = open(filename, "rb")
        pdfread = p2.PdfFileReader(pdf_file)
        if pdfread.isEncrypted:
            pdfread.decrypt(password)
        for i in range(0, pdfread.getNumPages()):
            pageinfo = pdfread.getPage(i)
            if pageinfo.extractText().casefold().find(word.casefold()) !=-1:
                lists = pageinfo.extractText().split("\n")
                for j in range(len(lists)):
                    if lists[j].casefold().count(word.casefold()) != 0:
                        #print("j-", lists[j])
                        list1 = lists[j].split(". ")
                        #print(len(list1))
                        mn = re.findall('[A-Z]',lists[j][0])
                        if len(list1) == 1 and len(mn)!= 0:
                            count = count + 1
                            ws['A' + str(count)] = filename.split("\\")[-1]
                            ws['B' + str(count)] = word
                            ws['C' + str(count)] = i + 1
                            ws['D' + str(count)] = list1[0]
                        if len(list1) == 2 and list1[1] == "":
                            str1 = ""
                            if len(mn)!= 0:
                                str1 = list1[0]
                            else:
                                for k in range(j-1,0,-1):
                                    try:
                                        if len(lists[k].split(". ")) >1:
                                            sent = lists[k].split(". ")[-1]
                                            ##print("dsent = ", len(sent), lists[k][-2:])
                                            if sent[-1] == "." or sent == '':
                                                ##print("hi")
                                                str1 = str1 + ""
                                                break
                                            else:
                                                str1 = sent + str1
                                                break
                                        elif len(lists[k].split(". ")) == 1 and lists[k].split(". ")[-1][-1]==".":
                                            str1 = lists[k]
                                            break
                                        else:
                                            str1 = lists[k] + str1
                                    except:
                                        pass
                                str1 = str1 + list1[0]
                            count = count + 1
                            ws['A' + str(count)] = filename.split("\\")[-1]
                            ws['B' + str(count)] = word
                            ws['C' + str(count)] = i + 1
                            ws['D' + str(count)] = str1
                            #print("d = ", lists[j])
                            #print(str1)
                        if j == len(lists)-1:
                            if len(list1) == 1 and len(mn)!= 0:
                                #print("c")
                                #print(lists[j])
                                count = count + 1
                                ws['A' + str(count)] = filename.split("\\")[-1]
                                ws['B' + str(count)] = word
                                ws['C' + str(count)] = i + 1
                                ws['D' + str(count)] = lists[j]
                                    
                        else:
                            try:
                                xy = re.findall('[A-Z]',lists[j+1][0])
                            except:
                                xy = " "
                            if len(list1) == 1 and len(mn)!=0 and len(xy)!=0:
                                #print("b")
                                #print(lists[j])
                                count = count + 1
                                ws['A' + str(count)] = filename.split("\\")[-1]
                                ws['B' + str(count)] = word
                                ws['C' + str(count)] = i + 1
                                ws['D' + str(count)] = lists[j]
                        
                        if len(list1) == 1 and len(mn)==0 :
                            #print("mn=", lists[j][0])
                            str1 = ""
                            xz = re.findall('[A-Za-z0-9]',lists[j][0])
                            if lists[j][0] == " " or len(xz) == 0:
                                str1 = lists[j]
                            else:
                                for k in range(j-1,0,-1):
                                    try:
                                        if len(lists[k].split(". ")) >1:
                                            sent = lists[k].split(". ")[-1]
                                            if sent[-1] == ".":
                                                str1 = str1 + ""
                                                break
                                            else:
                                                str1 = sent + str1
                                                break
                                        elif len(lists[k].split(". ")) == 1 and lists[k].split(". ")[-1][-1]==".":
                                            str1 = str1 + ""
                                            break
                                        else:
                                            str1 = lists[k] + str1
                                    except:
                                        pass
                                str1 = str1 + list1[0]
                                
                                for k in range(j+1,len(lists[j])):
                                    try:
                                        if len(lists[k].split(". ")) >1:
                                            sent = lists[k].split(". ")[0]
                                            str1 = str1 + sent
                                            break
                                        elif len(lists[k].split(". ")) == 1 and lists[k].split(". ")[0][-1] == "." :
                                            str1 = str1 + lists[k]
                                            break
                                        else:
                                            str1 = str1 + lists[k]
                                    except:
                                        pass
                                str1 = list1[-1] + str1
                                pl = re.findall('[a-zA-Z0-9]', lists[j-1])
                                if lists[j-1][-1] == "." or len(pl) == 0:
                                    str1 = lists[j]
                            #print("a=",str1)
                            count = count + 1
                            ws['A' + str(count)] = filename.split("\\")[-1]
                            ws['B' + str(count)] = word
                            ws['C' + str(count)] = i + 1
                            ws['D' + str(count)] = str1
                                                    
                        if len(list1)>1 and list1[1]!= "":
                            #if list1[0].casefold().find(word.casefold()) == -1 and list1[-1].casefold().find(word.casefold()) == -1:
                            if len(list1)>2:
                                for m in range(1,len(list1)-1):
                                    if list1[m].casefold().find(word.casefold()) != -1:
                                        #print("l=", list1[m])
                                        count = count + 1
                                        ws['A' + str(count)] = filename.split("\\")[-1]
                                        ws['B' + str(count)] = word
                                        ws['C' + str(count)] = i + 1
                                        ws['D' + str(count)] = list1[m]
                            if list1[0].casefold().find(word.casefold()) != -1:
                                str1 = ""
                                if len(mn)!=0:
                                    str1 = list1[0]
                                    #print("play", str1)
                                else:
                                    for k in range(j-1,0,-1):
                                        try:
                                            if len(lists[k].split(". ")) >1:
                                                sent = lists[k].split(". ")[-1]
                                                if sent[-1] == ".":
                                                        str1 = str1 + ""
                                                        break
                                                else:
                                                    str1 = sent + str1
                                                    break
                                            elif len(lists[k].split(". ")) == 1 and lists[k].split(". ")[-1][-1]==".":
                                                str1 = lists[k]
                                                break
                                            else:
                                                str1 = lists[k] + str1
                                        except:
                                            pass
                                    str1 = str1 + list1[0]
                                    #print("k=", str1)
                                count = count + 1
                                ws['A' + str(count)] = filename.split("\\")[-1]
                                ws['B' + str(count)] = word
                                ws['C' + str(count)] = i + 1
                                ws['D' + str(count)] = str1
                            
                            if list1[-1].casefold().find(word.casefold()) != -1:
                                str1 = ""
                                if list1[-1][-1] == ".":
                                    str1 = list1[-1]
                                else:
                                    for k in range(j+1,len(lists[j])):
                                        try:
                                            if len(lists[k].split(". ")) >1:
                                                sent = lists[k].split(". ")[0]
                                                str1 = str1 + sent
                                                break
                                            elif len(lists[k].split(". ")) == 1 and lists[k].split(". ")[0][-1] == "." :
                                                str1 = str1 + lists[k]
                                                break
                                            else:
                                                str1 = str1 + lists[k]
                                        except:
                                            pass
                                    str1 = list1[-1] + str1
                                #print("n=", str1)
                                count = count + 1
                                ws['A' + str(count)] = filename.split("\\")[-1]
                                ws['B' + str(count)] = word
                                ws['C' + str(count)] = i + 1
                                ws['D' + str(count)] = str1
                            
#                         if len(list1) == 1 and j!=0:
#                             mn = re.findall('A-Z', words[0])
#                             #print(len(mn))
#                             if len(lists[j-1].split(". ")) == 1 and lists[j-1][-1] != "." and len(mn) == 0:
#                                 wf = words.casefold().find(word.casefold())
#                                 try:
#                                     letr = re.findall("[a-zA-Z]", words[wf+len(word)])
#                                 except:
#                                     letr = ["1"]
#                                 if wf != -1 and len(letr) == 0 or len(words) == wf+len(word):
#                                     count = count + 1
#                                     ws['A' + str(count)] = filename.split("\\")[-1]
#                                     ws['B' + str(count)] = word
#                                     ws['C' + str(count)] = i + 1
#                                     #print("x-", lists[j-1] + words)
#                                     ws['D' + str(count)] = lists[j-1] + words
#                             elif len(lists[j-1].split(". ")) > 1 and lists[j-1].split(". ")[-1] != "" and len(mn) == 0:
#                                 wf = words.casefold().find(word.casefold())
#                                 try:
#                                     letr = re.findall("[a-zA-Z]", words[wf+len(word)])
#                                 except:
#                                     letr = ["1"]
#                                 if wf != -1 and len(letr) == 0 or len(words) == wf+len(word):
#                                     count = count + 1
#                                     ws['A' + str(count)] = filename.split("\\")[-1]
#                                     ws['B' + str(count)] = word
#                                     ws['C' + str(count)] = i + 1
#                                     #print("x-", lists[j-1] + words)
#                                     ws['D' + str(count)] = lists[j-1].split(". ")[-1] + words
#                                 
#                         if len(list1) == 1 or len(list1)==2 and len(list1[1]) == 0:
#                             mn = re.findall('A-Z', words[0])
#                             if j+1 != len(lists):
#                                 xy = re.findall('A-Z',lists[j+1].split(". ")[0][0])
#                             else:
#                                 xy = ""
#                             if len(mn) != 0 and len:
#                                 #print("y=", lists[j])
#                                 wf = words.casefold().find(word.casefold())
#                                 try:
#                                     letr = re.findall("[a-zA-Z]", words[wf+len(word)])
#                                 except:
#                                     letr = ["1"]
#                                 if wf != -1 and len(letr) == 0 or len(words) == wf+len(word):
#                                     count = count + 1
#                                     ws['A' + str(count)] = filename.split("\\")[-1]
#                                     ws['B' + str(count)] = word
#                                     ws['C' + str(count)] = i + 1
#                                     ws['D' + str(count)] = words
#                             else:
#                                 #print("y=", lists[j])
#                         else:
#                             if words != list1[0] or words!= list1[-1]:
#                                 wf = words.casefold().find(word.casefold())
#                                 try:
#                                     letr = re.findall("[a-zA-Z]", words[wf+len(word)])
#                                 except:
#                                     letr = ["1"]
#                                 if wf != -1 and len(letr) == 0 or len(words) == wf+len(word):
#                                     count = count + 1
#                                     ws['A' + str(count)] = filename.split("\\")[-1]
#                                     ws['B' + str(count)] = word
#                                     ws['C' + str(count)] = i + 1
#                                     ws['D' + str(count)] = words
#                             elif words == list1[0]:
#                                 try:
#                                     wf = words.casefold().find(word.casefold())
#                                     try:
#                                         letr = re.findall("[a-zA-Z]", words[wf+len(word)])
#                                     except:
#                                         letr = ["1"]
#                                     if wf != -1 and len(letr) == 0 or len(words) == wf+len(word):
#                                         count = count + 1
#                                         ws['A' + str(count)] = filename.split("\\")[-1]
#                                         ws['B' + str(count)] = word
#                                         ws['C' + str(count)] = i + 1
#                                         ws['D' + str(count)] = lists[j-1].split(". ")[-1] + words
#                                 except:
#                                     pass
#                             
#                             else:
#                                 try:
#                                     wf = words.casefold().find(word.casefold())
#                                     try:
#                                         letr = re.findall("[a-zA-Z]", words[wf+len(word)])
#                                     except:
#                                         letr = ["1"]
#                                     if wf != -1 and len(letr) == 0 or len(words) == wf+len(word):
#                                         count = count + 1
#                                         ws['A' + str(count)] = filename.split("\\")[-1]
#                                         ws['B' + str(count)] = word
#                                         ws['C' + str(count)] = i + 1
#                                         if words[-1] != ".":
#                                             ws['D' + str(count)] = words + lists[j+1][0]
#                                         else:
#                                             ws['D' + str(count)] = words
#                                 except:
#                                     pass
                                
        
        pdf_file.close()

for value in sheet.iter_rows(max_col=2, values_only=True):
    word = value[0]
    password = value[1]
    if password is not None:
        updating(word, password)
    else:
        updating(word)

wb.save("PDFextract.xlsx")
